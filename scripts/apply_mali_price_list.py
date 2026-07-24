"""
Match the 2026-07-22 Mali price list (营销/网站价格.xlsx) to products in the
shared catalog (data/catalog_2026/products.json) and patch matched entries'
price:0 to their real XOF price in Mali-website/assets/data/products.js —
same role as apply_price_list.py plays for Cameroon, but:
  - source sheet has a different shape (No./Nom/型号/网站价格/备注, one flat
    list, not the Cameroon sheet's per-category "Model" header blocks)
  - column D ("网站价格" = "website price") is the final XOF price already —
    no markup applied, confirmed 2026-07-24 (unlike Cameroon's +20%)
  - target file lives in the separate Mali-website git repo, not this one

Same "deliberately conservative" rule as apply_price_list.py: only price a
SKU when its spec maps to exactly one price-list value, else leave price:0
("Prix sur demande") rather than guess.

Solar panels are NOT priced by this script. The catalog's panel `wattage`
field is a broad range per SKU (e.g. "560-585W") with heavy overlap across
99 panels, and Mali's list only gives a single wattage number per row — a
dry-run check found >90% of Mali panel rows match 3+ overlapping catalog
SKUs by wattage-range containment, i.e. genuinely ambiguous, not a parsing
gap. Panels stay price:0 until a real per-SKU cross-reference exists.
Controllers, cables, connectors, and fans are excluded for the same reason
(catalog capacity_ah for controllers is a range like "2-60A"; cables/
connectors/fans have one generic catalog SKU per family with no gauge/color
attribute to disambiguate against Mali's per-variant rows).

Run: python3 scripts/apply_mali_price_list.py [--dry-run]
"""
import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from apply_price_list import (  # noqa: E402
    match_gel_agm,
    match_lithium_kwh,
    match_inverter_kw,
    _first_number,
    _voltage_tokens,
)

CATALOG_JSON = Path("/Users/lucasfeng/rest-solar-agent/data/catalog_2026/products.json")
SOURCE_XLSX = Path("/Users/lucasfeng/Downloads/马里网站价格含大部分型号.7.22.xlsx")
PRODUCTS_JS = Path("/Users/lucasfeng/Mali-website/assets/data/products.js")


def parse_sheet():
    """Yields (category, normalized_spec, price) tuples in the same shape
    apply_price_list.py's match_* functions expect, extracted from Mali's
    free-text "Nom du produit" column via regex instead of Cameroon's
    per-category sheet structure."""
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, price = row[1], row[3]
        if not name or not isinstance(price, (int, float)):
            continue
        name_u = str(name).upper()

        if "GEL BATTERY" in name_u or "AGM" in name_u or "BATTERIE AU GEL" in name_u:
            m = re.search(r"(\d+(?:\.\d+)?)\s*V\s*(\d+(?:\.\d+)?)\s*AH?\b", name_u)
            if not m:
                continue
            subcat = "GEL battery" if "GEL" in name_u else "AGM battery"
            yield subcat, f"{m.group(1)}V*{m.group(2)}AH", round(price)
        elif "LITHIUM BATTERY" in name_u:
            m = re.search(r"([\d.]+)\s*KWH", name_u)
            if not m:
                continue
            yield "Lithium Battery", f"{m.group(1)}KW", round(price)
        elif "ONDULEUR" in name_u or ("INVERTER" in name_u and "HYBRID" in name_u):
            m = re.match(r"^[^\d]*([\d.]+)\s*KW\s*$", str(name).strip(), re.I)
            if not m:
                continue
            yield "Inverter", f"{m.group(1)}KW", round(price)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    sheet_rows = list(parse_sheet())
    gel_prices = match_gel_agm(sheet_rows, "GEL battery")
    agm_prices = match_gel_agm(sheet_rows, "AGM battery")
    lithium_prices = match_lithium_kwh(sheet_rows)
    inverter_prices = match_inverter_kw(sheet_rows)

    raw_products = json.loads(CATALOG_JSON.read_text(encoding="utf-8"))

    updates: dict[str, tuple[int, str]] = {}
    for p in raw_products:
        sku = p["sku"]
        cat = p["category"]
        sub = p.get("subcategory") or ""
        price = None
        reason = ""
        if cat == "batteries" and sub == "GEL Battery":
            for vtok in _voltage_tokens(p.get("voltage")):
                ah = _first_number((p.get("capacity_ah") or "").replace("Ah", "").replace("AH", ""))
                key = (vtok.upper(), ah)
                if key in gel_prices:
                    price, reason = gel_prices[key], f"GEL {key}"
                    break
        elif cat == "batteries" and sub == "AGM Battery":
            for vtok in _voltage_tokens(p.get("voltage")):
                ah = _first_number((p.get("capacity_ah") or "").replace("Ah", "").replace("AH", ""))
                key = (vtok.upper(), ah)
                if key in agm_prices:
                    price, reason = agm_prices[key], f"AGM {key}"
                    break
        elif cat == "batteries" and sub == "LiFePO4 Battery":
            kwh = _first_number((p.get("capacity_kwh") or "").replace("kWh", "").replace("KWh", "").replace("KW", ""))
            if kwh in lithium_prices:
                price, reason = lithium_prices[kwh], f"Li {kwh}kWh"
        elif cat == "inverters":
            kw = _first_number((p.get("power_kw") or "").replace("kW", "").replace("KW", ""))
            if kw in inverter_prices:
                price, reason = inverter_prices[kw], f"Inverter {kw}kW"

        if price is not None:
            updates[sku] = (price, f"{sku} {p.get('model')!r} <- {reason}")

    print(f"Parsed {len(sheet_rows)} matchable-category rows from {SOURCE_XLSX.name}.")
    print(f"Matched {len(updates)} of {len(raw_products)} products:\n")
    for sku, (price, reason) in updates.items():
        print(f"  {price:>8,} XOF  {reason}")

    if args.dry_run:
        print("\n--dry-run: no changes written.")
        return

    js = PRODUCTS_JS.read_text(encoding="utf-8")
    applied = 0
    for sku, (price, _reason) in updates.items():
        pattern = re.compile(rf'(id:"{re.escape(sku)}"(?:(?!price:)[^}}])*?price:)0(,)')
        new_js, n = pattern.subn(rf"\g<1>{price}\g<2>", js)
        if n == 1:
            js = new_js
            applied += 1
        elif n == 0:
            print(f"  WARNING: could not find price:0 for {sku} in products.js")
        else:
            print(f"  WARNING: {sku} matched {n} times in products.js, skipped")

    PRODUCTS_JS.write_text(js, encoding="utf-8")
    print(f"\nApplied {applied} price updates to {PRODUCTS_JS}.")


if __name__ == "__main__":
    main()
